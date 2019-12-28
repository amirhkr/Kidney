from openpyxl import load_workbook
import os
from datetime import datetime
import dateutil.relativedelta

# load workbooks
wb_KFRE = load_workbook('C:/Users/s4597917/Desktop/Kidney/Amir_KFRE_MetroNorth_UPDATED2.xlsx')
wb_PathologyMaster = load_workbook('C:/Users/s4597917/Desktop/Kidney/Path3.xlsx')

sheet_KFRE = wb_KFRE.active
sheet_PathologyMaster = wb_PathologyMaster.active
tempArray = []
myDict = {}

for row_KFRE in sheet_KFRE.iter_rows(min_row=3, max_col=sheet_KFRE.max_column, max_row=sheet_KFRE.max_row):
    tempArray.clear()
    for rowPath in sheet_PathologyMaster.iter_rows(min_row=2, max_col=sheet_PathologyMaster.max_column, max_row=sheet_PathologyMaster.max_row):
        if row_KFRE[0].value == rowPath[0].value:
            # myDict.clear()
            myDict = {
                "PatientId": rowPath[0].value,
                "datePath": rowPath[1].value,
                "variableName": rowPath[2].value,
                "variableValue": rowPath[3].value,
                "unit": rowPath[4].value
            }
            tempArray.append(myDict)

    date_KFRE = row_KFRE[9].value
    date_KFRE_18MonthPast = date_KFRE - dateutil.relativedelta.relativedelta(months=18)
    date_KFRE_21MonthPast = date_KFRE - dateutil.relativedelta.relativedelta(months=21)
    date_KFRE_24MonthPast = date_KFRE - dateutil.relativedelta.relativedelta(months=24)
    date_KFRE_27MonthPast = date_KFRE - dateutil.relativedelta.relativedelta(months=27)
    date_KFRE_30MonthPast = date_KFRE - dateutil.relativedelta.relativedelta(months=30)

    GFR_Array = []
    GFR_iSTAT_Array = []
    URINE_Array = []
    dictTEST = {}

    #  3 month range
    for i in range(len(tempArray)):
        if(((tempArray[i]['datePath'] <= date_KFRE_21MonthPast) and (
                tempArray[i]['datePath'] >= date_KFRE_24MonthPast)) or
                ((tempArray[i]['datePath'] >= date_KFRE_27MonthPast) and (
                        tempArray[i]['datePath'] <= date_KFRE_24MonthPast))):
            dictTEST = {
                    "pl2_date": tempArray[i]['datePath'],
                    "pl2_name": tempArray[i]['variableName'],
                    "pl2_value": tempArray[i]['variableValue'],
                    "pl2_unit": tempArray[i]['unit']
                 }

            if (tempArray[i]['variableName'] == 'GFR') or (tempArray[i]['variableName'] == 'eGFR') or (tempArray[i]['variableName'] == 'GFR (estimated)'):
                GFR_Array.append(dictTEST)

            if tempArray[i]['variableName'] == 'GFR (estimated) iSTAT':
                GFR_iSTAT_Array.append(dictTEST)

    #  6 month range
    if len(GFR_Array) == 0 and len(GFR_iSTAT_Array) == 0:# no variable found so keep searching for +/-6 month span
        for i in range(len(tempArray)):
            if (((tempArray[i]['datePath'] <= date_KFRE_18MonthPast) and (
                    tempArray[i]['datePath'] >= date_KFRE_24MonthPast)) or
                    ((tempArray[i]['datePath'] >= date_KFRE_30MonthPast) and (
                            tempArray[i]['datePath'] <= date_KFRE_24MonthPast))):
                dictTEST = {
                    "pl2_date": tempArray[i]['datePath'],
                    "pl2_name": tempArray[i]['variableName'],
                    "pl2_value": tempArray[i]['variableValue'],
                    "pl2_unit": tempArray[i]['unit']
                }
                if (tempArray[i]['variableName'] == 'GFR') or (tempArray[i]['variableName'] == 'eGFR') or (tempArray[i]['variableName'] == 'GFR (estimated)'):
                    GFR_Array.append(dictTEST)

                if tempArray[i]['variableName'] == 'GFR (estimated) iSTAT':
                    GFR_iSTAT_Array.append(dictTEST)

    # it's time to use the actual results
    resultGFR_Dict = {}
    if len(GFR_Array) == 0:
        if len(GFR_iSTAT_Array) == 1:
            resultGFR_Dict = GFR_iSTAT_Array[0]

        if len(GFR_iSTAT_Array) >= 2:
            for i in range(len(GFR_iSTAT_Array)):
                if i <= len(GFR_iSTAT_Array)-2:
                    date1 = GFR_iSTAT_Array[i]['pl2_date']
                    date2 = GFR_iSTAT_Array[i+1]['pl2_date']
                   #  older than 24
                    if (date1 >= date_KFRE_24MonthPast) and (date2 >= date_KFRE_24MonthPast):
                       if date1 >= date2:
                           resultGFR_Dict = GFR_iSTAT_Array[i]
                       else:
                           resultGFR_Dict = GFR_iSTAT_Array[i + 1]
                    #  younger than 24
                    if date1 < date_KFRE_24MonthPast and date2 < date_KFRE_24MonthPast:
                       if date1 < date2:
                           resultGFR_Dict = GFR_iSTAT_Array[i + 1]
                       else:
                           resultGFR_Dict = GFR_iSTAT_Array[i]
                    #  one right, one left
                    if date1 > date_KFRE_24MonthPast > date2:
                        if (date1 - date_KFRE_24MonthPast) <= (date_KFRE_24MonthPast - date2):# equal distance
                            resultGFR_Dict = GFR_iSTAT_Array[i]
                        else:
                            resultGFR_Dict = GFR_iSTAT_Array[i + 1]
                    #  one left, one right
                    if date1 < date_KFRE_24MonthPast < date2:
                        if (date_KFRE_24MonthPast - date1) <= (date2 - date_KFRE_24MonthPast):# equal distance
                            resultGFR_Dict = GFR_iSTAT_Array[i]
                        else:
                            resultGFR_Dict = GFR_iSTAT_Array[i + 1]
                    GFR_iSTAT_Array[i+1] = resultGFR_Dict
    else:
        if len(GFR_Array) == 1:
            resultGFR_Dict = GFR_Array[0]
        if len(GFR_Array) >= 2:
            for i in range(len(GFR_Array)):
                if i <= len(GFR_Array)-2:
                    date1 = GFR_Array[i]['pl2_date']
                    date2 = GFR_Array[i+1]['pl2_date']
                   #  older than 24
                    if (date1 >= date_KFRE_24MonthPast) and (date2 >= date_KFRE_24MonthPast):
                       if date1 >= date2:
                           resultGFR_Dict = GFR_Array[i]
                       else:
                           resultGFR_Dict = GFR_Array[i + 1]
                    #  younger than 24
                    if date1 < date_KFRE_24MonthPast and date2 < date_KFRE_24MonthPast:
                       if date1 < date2:
                           resultGFR_Dict = GFR_Array[i + 1]
                       else:
                           resultGFR_Dict = GFR_Array[i]
                    #  one right, one left
                    if date1 > date_KFRE_24MonthPast > date2:
                        if (date1 - date_KFRE_24MonthPast) <= (date_KFRE_24MonthPast - date2):# equal distance
                            resultGFR_Dict = GFR_Array[i]
                        else:
                            resultGFR_Dict = GFR_Array[i + 1]
                    #  one left, one right
                    if date1 < date_KFRE_24MonthPast < date2:
                        if (date_KFRE_24MonthPast - date1) <= (date2 - date_KFRE_24MonthPast):# equal distance
                            resultGFR_Dict = GFR_Array[i]
                        else:
                            resultGFR_Dict = GFR_Array[i + 1]
                    GFR_Array[i+1] = resultGFR_Dict  # so we need to update the next element with the current result
                    # because we want it ti be part of comparison for the next iteration

    #  update Excell
#     row_KFRE[10].value = resultGFR_Dict['pl2_date']
#     row_KFRE[11].value = resultGFR_Dict['pl2_value']
#
# wb_KFRE.save()
wb_KFRE.close()
wb_PathologyMaster.close()