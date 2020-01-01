import dateutil.relativedelta
from openpyxl import load_workbook
import datetime

# load workbooks
KFR_file_Path = 'C:/Users/s4597917/Desktop/Kidney/Amir_KFRE_MetroNorth_UPDATED3.xlsx'
Pathology_file_path = 'C:/Users/s4597917/Desktop/Kidney/pathology-MASTER3.xlsx'

# KFR_file_Path = 'C:/Users/s4597917/Desktop/Kidney/Amir_KFRE2.xlsx'
# Pathology_file_path = 'C:/Users/s4597917/Desktop/Kidney/Path2.xlsx'

wb_KFRE = load_workbook(KFR_file_Path)
wb_PathologyMaster = load_workbook(Pathology_file_path)

sheet_KFRE = wb_KFRE.active
sheet_PathologyMaster = wb_PathologyMaster.active
RequiredTestNames_GFR = [
                    'GFR',
                    'eGFR',
                    'GFR (estimated)',
                    'GFR (estimated) iSTAT']

RequiredTestNames_URINE = [
                    'Albumin/Creatinine ratio',
                    'R U-Albumin/Creat',
                    'Protein/Creatinine',
                    'R-U-Protein/Creat',
                    'Urine Albumin',
                    'Calculated U-Albumin Excretion',
                    'Urine Total protein']

RequiredTestNames_URINE_ACR = [
                    'Albumin/Creatinine ratio',
                    'R U-Albumin/Creat',
                    'Urine Albumin',
                    'Calculated U-Albumin Excretion']

RequiredTestNames_URINE_PCR = [
                    'Protein/Creatinine',
                    'R-U-Protein/Creat',
                    'Urine Total protein']

PathologyArray = []
tempArray = []
myDict = {}
KFRE_i = 0
Path_i = 0

# read pathology file ONLY ONE TIME and store it in the array in memory and we go from there
for rowPath in sheet_PathologyMaster.iter_rows(min_row=2, max_col=sheet_PathologyMaster.max_column,
                                               max_row=sheet_PathologyMaster.max_row):
    Path_i = Path_i + 1
    # print("Processing: " + str(Path_i) + " of " + str(sheet_PathologyMaster.max_row - 2) + "...")
    print("Processing: " + str(Path_i))

    # myDict.clear()
    if rowPath[2].value in RequiredTestNames_GFR or rowPath[2].value in RequiredTestNames_URINE:
        PathologyArray.append({
        "PatientId": rowPath[0].value,
        "datePath": rowPath[1].value,
        "variableName": rowPath[2].value,
        "variableValue": rowPath[3].value,
        "unit": rowPath[4].value
        })
# Completed

print('Completed reading pathology file!')

for row_KFRE in sheet_KFRE.iter_rows(min_row=3, max_col=sheet_KFRE.max_column, max_row=sheet_KFRE.max_row):
    tempArray.clear()
    resultGFR_Dict = {}
    resultURINE = {}
    KFRE_i = KFRE_i + 1
    for rowPath in PathologyArray:
        Path_i = Path_i + 1
        print("Processing: "+str(KFRE_i)+" of "+str(sheet_KFRE.max_row-2)+" --- "+row_KFRE[0].value+"  "+rowPath['PatientId'])
        if row_KFRE[0].value == rowPath['PatientId']:
            # myDict.clear()
            myDict = {
                "PatientId": rowPath['PatientId'],
                "datePath": rowPath['datePath'],
                "variableName": rowPath['variableName'],
                "variableValue": rowPath['variableValue'],
                "unit": rowPath['unit']
            }
            tempArray.append(myDict)

    date_KFRE = row_KFRE[9].value
    date_KFRE_18MonthPast = date_KFRE - dateutil.relativedelta.relativedelta(months=18)
    date_KFRE_21MonthPast = date_KFRE - dateutil.relativedelta.relativedelta(months=21)
    date_KFRE_24MonthPast = date_KFRE - dateutil.relativedelta.relativedelta(months=24)
    date_KFRE_27MonthPast = date_KFRE - dateutil.relativedelta.relativedelta(months=27)
    date_KFRE_30MonthPast = date_KFRE - dateutil.relativedelta.relativedelta(months=30)

    GFR_Array = []
    GFR_iSTAT_Array = []
    URINE_Array = []
    dictTEST = {}

    #  3 month range
    for i in range(len(tempArray)):
        if(((tempArray[i]['datePath'] <= date_KFRE_21MonthPast) and (
                tempArray[i]['datePath'] >= date_KFRE_24MonthPast)) or
                ((tempArray[i]['datePath'] >= date_KFRE_27MonthPast) and (
                        tempArray[i]['datePath'] <= date_KFRE_24MonthPast))):
            dictTEST = {
                    "pl2_date": tempArray[i]['datePath'],
                    "pl2_name": tempArray[i]['variableName'],
                    "pl2_value": tempArray[i]['variableValue'],
                    "pl2_unit": tempArray[i]['unit']
                 }

            if (tempArray[i]['variableName'] == 'GFR') or (tempArray[i]['variableName'] == 'eGFR') or (tempArray[i]['variableName'] == 'GFR (estimated)'):
                GFR_Array.append(dictTEST)

            if tempArray[i]['variableName'] == 'GFR (estimated) iSTAT':
                GFR_iSTAT_Array.append(dictTEST)

            if tempArray[i]['variableName'] in RequiredTestNames_URINE:
                URINE_Array.append(dictTEST)

    #  6 month range
    if (len(GFR_Array) == 0 and len(GFR_iSTAT_Array) == 0) or (len(URINE_Array) == 0):# no variable found so keep searching for +/-6 month span
        for i in range(len(tempArray)):
            if (((tempArray[i]['datePath'] <= date_KFRE_18MonthPast) and (
                    tempArray[i]['datePath'] >= date_KFRE_24MonthPast)) or
                    ((tempArray[i]['datePath'] >= date_KFRE_30MonthPast) and (
                            tempArray[i]['datePath'] <= date_KFRE_24MonthPast))):
                dictTEST = {
                    "pl2_date": tempArray[i]['datePath'],
                    "pl2_name": tempArray[i]['variableName'],
                    "pl2_value": tempArray[i]['variableValue'],
                    "pl2_unit": tempArray[i]['unit']
                }
                if (tempArray[i]['variableName'] == 'GFR') or (tempArray[i]['variableName'] == 'eGFR') or (tempArray[i]['variableName'] == 'GFR (estimated)'):
                    GFR_Array.append(dictTEST)

                if tempArray[i]['variableName'] == 'GFR (estimated) iSTAT':
                    GFR_iSTAT_Array.append(dictTEST)

                if tempArray[i]['variableName'] in RequiredTestNames_URINE:
                    URINE_Array.append(dictTEST)

    # it's time to use the actual results
    # resultGFR_Dict = {}
    if len(GFR_Array) == 0:
        if len(GFR_iSTAT_Array) == 1:
            resultGFR_Dict = GFR_iSTAT_Array[0]

        if len(GFR_iSTAT_Array) >= 2:
            for i in range(len(GFR_iSTAT_Array)):
                if i <= len(GFR_iSTAT_Array)-2:
                    date1 = GFR_iSTAT_Array[i]['pl2_date']
                    date2 = GFR_iSTAT_Array[i+1]['pl2_date']
                   #  older than 24
                    if (date1 >= date_KFRE_24MonthPast) and (date2 >= date_KFRE_24MonthPast):
                       if date1 >= date2:
                           resultGFR_Dict = GFR_iSTAT_Array[i]
                       else:
                           resultGFR_Dict = GFR_iSTAT_Array[i + 1]
                    #  younger than 24
                    if date1 < date_KFRE_24MonthPast and date2 < date_KFRE_24MonthPast:
                       if date1 < date2:
                           resultGFR_Dict = GFR_iSTAT_Array[i + 1]
                       else:
                           resultGFR_Dict = GFR_iSTAT_Array[i]
                    #  one right, one left
                    if date1 > date_KFRE_24MonthPast > date2:
                        if (date1 - date_KFRE_24MonthPast) <= (date_KFRE_24MonthPast - date2):# equal distance
                            resultGFR_Dict = GFR_iSTAT_Array[i]
                        else:
                            resultGFR_Dict = GFR_iSTAT_Array[i + 1]
                    #  one left, one right
                    if date1 < date_KFRE_24MonthPast < date2:
                        if (date_KFRE_24MonthPast - date1) <= (date2 - date_KFRE_24MonthPast):# equal distance
                            resultGFR_Dict = GFR_iSTAT_Array[i]
                        else:
                            resultGFR_Dict = GFR_iSTAT_Array[i + 1]
                    GFR_iSTAT_Array[i+1] = resultGFR_Dict
    else:
        if len(GFR_Array) == 1:
            resultGFR_Dict = GFR_Array[0]
        if len(GFR_Array) >= 2:
            for i in range(len(GFR_Array)):
                if i <= len(GFR_Array)-2:
                    date1 = GFR_Array[i]['pl2_date']
                    date2 = GFR_Array[i+1]['pl2_date']
                   #  older than 24
                    if (date1 >= date_KFRE_24MonthPast) and (date2 >= date_KFRE_24MonthPast):
                       if date1 >= date2:
                           resultGFR_Dict = GFR_Array[i]
                       else:
                           resultGFR_Dict = GFR_Array[i + 1]
                    #  younger than 24
                    if date1 < date_KFRE_24MonthPast and date2 < date_KFRE_24MonthPast:
                       if date1 < date2:
                           resultGFR_Dict = GFR_Array[i + 1]
                       else:
                           resultGFR_Dict = GFR_Array[i]
                    #  one right, one left
                    if date1 > date_KFRE_24MonthPast > date2:
                        if (date1 - date_KFRE_24MonthPast) <= (date_KFRE_24MonthPast - date2):# equal distance
                            resultGFR_Dict = GFR_Array[i]
                        else:
                            resultGFR_Dict = GFR_Array[i + 1]
                    #  one left, one right
                    if date1 < date_KFRE_24MonthPast < date2:
                        if (date_KFRE_24MonthPast - date1) <= (date2 - date_KFRE_24MonthPast):# equal distance
                            resultGFR_Dict = GFR_Array[i]
                        else:
                            resultGFR_Dict = GFR_Array[i + 1]
                    GFR_Array[i+1] = resultGFR_Dict  # so we need to update the next element with the current result
                    # because we want it ti be part of comparison for the next iteration

    #  update Excell
    if 'pl2_date' in resultGFR_Dict:
        row_KFRE[10].value = str(resultGFR_Dict['pl2_date'].day)+"/"+str(resultGFR_Dict['pl2_date'].month)+"/"+str(resultGFR_Dict['pl2_date'].year)
    else:
        row_KFRE[10].value = ''

    if 'pl2_value' in resultGFR_Dict:
        row_KFRE[11].value = resultGFR_Dict['pl2_value']
    else:
        row_KFRE[11].value = ''


    # --------------hanld urine array-------------
    if len(URINE_Array) == 0:
        resultURINE = {}
    if len(URINE_Array) == 1:
        resultURINE = URINE_Array[0]
    if len(URINE_Array) >= 2:
        for URINE in RequiredTestNames_URINE:
            OrderArray = []
            for k in range(len(URINE_Array)):
                if URINE == URINE_Array[k]['pl2_name']:
                    OrderArray.append(URINE_Array[k])

            if len(OrderArray) == 1:
                resultURINE = OrderArray[0]
            if len(OrderArray) >= 2:
                for i in range(len(OrderArray)):
                    if i <= len(OrderArray) - 2:
                        date1 = OrderArray[i]['pl2_date']
                        date2 = OrderArray[i + 1]['pl2_date']
                        #  older than 24
                        if (date1 >= date_KFRE_24MonthPast) and (date2 >= date_KFRE_24MonthPast):
                            if date1 >= date2:
                                resultURINE = OrderArray[i]
                            else:
                                resultURINE = OrderArray[i + 1]
                        #  younger than 24
                        if date1 < date_KFRE_24MonthPast and date2 < date_KFRE_24MonthPast:
                            if date1 < date2:
                                resultURINE = OrderArray[i + 1]
                            else:
                                resultURINE = OrderArray[i]
                        #  one right, one left
                        if date1 > date_KFRE_24MonthPast > date2:
                            if (date1 - date_KFRE_24MonthPast) <= (date_KFRE_24MonthPast - date2):  # equal distance
                                resultURINE = OrderArray[i]
                            else:
                                resultURINE = OrderArray[i + 1]
                        #  one left, one right
                        if date1 < date_KFRE_24MonthPast < date2:
                            if (date_KFRE_24MonthPast - date1) <= (date2 - date_KFRE_24MonthPast):  # equal distance
                                resultURINE = OrderArray[i]
                            else:
                                resultURINE = OrderArray[i + 1]
                        OrderArray[
                            i + 1] = resultURINE  # so we need to update the next element with the current result
                        # because we want it ti be part of comparison for the next iteration
            if 'pl2_date' in resultURINE:
                break

    #  update Excell
    if 'pl2_date' in resultURINE:
        row_KFRE[12].value = str(resultURINE['pl2_date'].day)+"/"+str(resultURINE['pl2_date'].month)+"/"+str(resultURINE['pl2_date'].year)
    else:
        row_KFRE[12].value = ''

    if 'pl2_name' in resultURINE:
        if resultURINE['pl2_name'] in RequiredTestNames_URINE_ACR:
            row_KFRE[13].value = resultURINE['pl2_name']
            row_KFRE[15].value = resultURINE['pl2_value']
            row_KFRE[16].value = resultURINE['pl2_unit']
        else:
            if resultURINE['pl2_name'] in RequiredTestNames_URINE_PCR:
                row_KFRE[14].value = resultURINE['pl2_name']
                row_KFRE[15].value = resultURINE['pl2_value']
                row_KFRE[16].value = resultURINE['pl2_unit']


wb_KFRE.save(KFR_file_Path)
wb_KFRE.close()
wb_PathologyMaster.close()
PathologyArray.clear()
print('Done!')